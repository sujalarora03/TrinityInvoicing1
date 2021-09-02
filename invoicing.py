import glob
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Fill
from num2words import num2words
import os
from win32com import client
from pywintypes import com_error
import time

xlsxfile = glob.glob('Test.xlsx')
print(xlsxfile)

wb = load_workbook(filename=f'{xlsxfile[0]}')

ws = wb.worksheets[0]
#print(ws)
ws1 = wb.worksheets[1]
#print(ws1)

loop_this = [x.value for x in ws1['B'] if x.value]
for k in range(3,len(loop_this)+2):
    print(f'K IS {k}')
    wb = load_workbook(filename=f'{xlsxfile[0]}')

    ws = wb.worksheets[0]
    #print(ws)
    ws1 = wb.worksheets[1]
    #print(ws1)
    
    #Full Name
    ws['B1'] = ws1[f'D{k}'].value.upper()
    print('Full Name = '+ws['B1'].value)
    ws['E40'] = 'For ' + ws1[f'D{k}'].value.upper()

    #Address
    ws['B2'] = 'Address: '+ ws1[f'O{k}'].value
    print(ws['B2'].value)

    #PAN
    ws['B5'] = 'PAN: ' + ws1[f'L{k}'].value.upper()
    print(ws['B5'].value)

    #Invoice No & Date
    ws['E6'] = f"INVOICE NO : {str(int(ws1[f'R{k}'].value)+1)} / 21-22 {chr(10)}DATE : {str(ws1['A1'].value).replace(' 00:00:00','')}"
    print(ws['E6'].value)
    ws1[f'R{k}'] = int(ws1[f'R{k}'].value) + 1
    print('No of Invoices Generated = '+ str(ws1[f'R{k}'].value))

    #Particulars
    ws['B10'] = f"'Professional services rendered by the individual as per the agreement'{chr(10)}Project Name : Facebook Partner Program{' - VOD' if ws1[f'A{k}'].value == 'VOD' else ''}"
    print('Particulars = '+ ws['B10'].value)

    #Hours
    ws['D10'] = ws1[f'N{k}'].value
    print('Hours = '+ str(ws['D10'].value))

    #Amount
    ws['F10'] = ws1[f'Q{k}'].value
    print('Amount = '+ str(ws['F10'].value))

    #GrandTotal
    if '.' in str(ws1[f'Q{k}'].value):
        rupees, paise = str(ws1[f'Q{k}'].value).split('.')
        if rupees:
            rupees = 'Total Amount (INR - In Words):    RUPEES ' + num2words(int(rupees), lang='en_IN').upper()
        if paise:
            paise = num2words(int(paise), lang='en_IN').upper() + ' PAISE ONLY'
            if rupees:
                paise = ' AND ' + paise
            ws['B26'] = rupees + paise
    else:
        rupees = str(ws1[f'Q{k}'].value)
        if rupees:
            rupees = 'Total Amount (INR - In Words):    RUPEES ' + num2words(int(rupees), lang='en_IN').upper() + ' ONLY'
        ws['B26'] = rupees
    print('Grand Total = '+ws['B26'].value)

    #Account Holder Name
    ws['B30'] = 'Account Holder Name: ' + ws1[f'E{k}'].value.upper()
    print(ws['B30'].value)

    #Bank Name
    ws['B31'] = 'Bank Name: ' + ws1[f'G{k}'].value.upper()
    print(ws['B31'].value)

    #Account Type
    ws['B32'] = 'Account Type: ' + ws1[f'I{k}'].value.upper()
    print(ws['B32'].value)

    #Account Number
    ws['B33'] = 'Bank Account Number: ' + str(ws1[f'F{k}'].value)
    print(ws['B33'].value)

    #Branch Address
    ws['B34'] = 'Branch: ' + ws1[f'H{k}'].value.upper()
    print(ws['B34'].value)

    #IFSC Code
    ws['B35'] = 'IFSC Code: ' + ws1[f'J{k}'].value.upper()
    print(ws['B35'].value)


    wb.save(f'{xlsxfile[0]}')


    openpyxl.worksheet.properties.PageSetupProperties(autoPageBreaks=None, fitToPage=True)


    if os.path.exists(str(os.getcwd())+'\\PDF'):
        pass
    else:
        os.makedirs(str(os.getcwd())+'\\PDF')
    # Path to original excel file
    WB_PATH = os.getcwd()+'\\Test.xlsx'
    # PDF path when saving
    PATH_TO_PDF = os.getcwd()+f"\\PDF\\\{ws1[f'D{k}'].value.upper()}.pdf"

    xlApp = client.Dispatch("Excel.Application")
    xlApp.Visible = False

    try:
        print('Start conversion to PDF')

        # Open
        wb = xlApp.Workbooks.Open(WB_PATH)
        print(wb)
        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = [1]
        wb.WorkSheets(ws_index_list).Select()

        # Save
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF, IncludeDocProperties=True)
    except com_error as e:
        print('failed.')
    else:
        print('Succeeded.')
        xlApp.Quit()
        del xlApp

